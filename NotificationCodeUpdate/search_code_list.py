# -*- coding: utf-8 -*-
"""
Created on Sat Jul  6 17:43:01 2019

@author: siwan
"""
"""
from openpyxl import load_workbook
import pandas as pd

excel_writer = pd.ExcelWriter('excel_trace.xlsx', engine='openpyxl')
wb = excel_writer.book

result = pd.DataFrame()
#excel_files =  ['Freezer_20190416.xlsx', 'Freezer_20190418.xlsx', 'Freezer_20190424.xlsx']
excel_files =  'Alex_20190604.xlsm'

df = pd.read_excel(excel_files)
df1 = df[['Code', 'Inward', 'PickUp Qty', 'New Balance', 'Pick up Details']]
"""

import time
from linebot import (LineBotApi, WebhookHandler)
from linebot.exceptions import (InvalidSignatureError)
from linebot.models import (MessageEvent, TextMessage, TextSendMessage,)
from datetime import datetime
from openpyxl import load_workbook

line_bot_api = LineBotApi('LqzzZREWJQl3tZ80owftCeC6UFzgW47s4Y/ZgFEj5SANQPC8s8m6XPZRNAAaeboSOofX1PUhP8Bby9+yZciQW3A9rajXAkjYHIf9dw3VK5peiAxrqd6OdMAOjsf5zeJZlmKI4L0/URAYsNY0liiKCgdB04t89/1O/w1cDnyilFU=')
handler = WebhookHandler('1bedc1daf1ea4c0119a4b9b86f46b5af')
#line_bot_api.push_message('U7965381d1b871ed0f2ac9a06c1a9f88e', TextSendMessage(text='has been updated'))
# refer to below to access remote
#f = pandas.read_excel(open('//X/str/Db/C/Source/selection/Date/Test/12.xlsx','rb'), sheetname='Sheet 1')

wb = load_workbook(filename=r"\\192.168.20.50\AlexServer\輸入共有\輸入共有フォルダー\SF Product Name & Code List(商品名確認票）\SF Product Name & Code List (商品名確認表)_for_test.xlsx", data_only=True)
sheets = wb.sheetnames
ws0 = wb[sheets[0]]

for row in ws0.iter_rows(min_row=1, max_col=10, max_row=ws0.max_row, values_only=True):

    if row[0] == 'N':
        print (row[0])
        line_bot_api.push_message('U7965381d1b871ed0f2ac9a06c1a9f88e', TextSendMessage(text=row[8] + ' has been added'))
    elif row[0] == 'U':
        print (row[0])
        line_bot_api.push_message('U7965381d1b871ed0f2ac9a06c1a9f88e', TextSendMessage(text=row[8] + ' has been updated'))

ws1 = wb[sheets[1]]

for row in ws1.iter_rows(min_row=1, max_col=10, max_row=ws1.max_row, values_only=True):

    if row[0] == 'N':
        print (row[0])
        line_bot_api.push_message('U7965381d1b871ed0f2ac9a06c1a9f88e', TextSendMessage(text=row[1] + ' has been added'))
    elif row[0] == 'U':
        print (row[0])
        line_bot_api.push_message('U7965381d1b871ed0f2ac9a06c1a9f88e', TextSendMessage(text=row[8] + ' has been updated'))
