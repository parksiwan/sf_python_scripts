
from openpyxl import load_workbook
import xlrd
import numpy as np
import pandas as pd


sf_orderform = 'SF_orderform.xlsx'
sf_code = 'SF_code.xlsx'
sf_code_local = 'SF_code_local.xlsx'

wb_code = xlrd.open_workbook(sf_code)
sheet_code = wb_code.sheet_by_index(0) 
code_list = []
for row_code in range (1, sheet_code.nrows):    
    code_data = {'sf_code' : sheet_code.cell(row_code, 8).value, 'sf_code_price' : sheet_code.cell(row_code, 99) }                
    code_list.append(code_data)
df_code = pd.DataFrame(code_list)    

wb_code_local = xlrd.open_workbook(sf_code_local)
sheet_code_local = wb_code_local.sheet_by_index(0)  
code_local_list = []
for row_code_local in range (1, sheet_code_local.nrows):    
    code_local_data = {'sf_code_local' : sheet_code_local.cell(row_code_local, 0).value, 
                      'sf_code_local_price' : sheet_code_local.cell(row_code_local, 9).value }                
    code_local_list.append(code_local_data)
df_code_local = pd.DataFrame(code_local_list)      

wb_orderform = xlrd.open_workbook(sf_orderform)
sheet_orderform = wb_orderform.sheet_by_index(0) 

temp_code = pd.DataFrame()
temp_code_local = pd.DataFrame()
df_result = pd.DataFrame()
for row_orderform in range (1, sheet_orderform.nrows):  
    temp_code = df_code[df_code['sf_code'].str.upper() == sheet_orderform.cell(row_orderform, 1).value]
    if len(temp_code) > 0:
        temp_code['orderform_price'] = sheet_orderform.cell(row_orderform, 4).value
        df_result = pd.concat([df_result, temp_code], ignore_index=True)
    else:
        temp_code_local = df_code_local[df_code_local['sf_code_local'].str.upper() == sheet_orderform.cell(row_orderform, 1).value]
        if len(temp_code_local) > 0:
            temp_code_local['orderform_price'] = sheet_orderform.cell(row_orderform, 4).value
            df_result = pd.concat([df_result, temp_code_local], ignore_index=True)
 
df_result.to_excel('MergeOrderformCodelist.xlsx', sheet_name='sheet1')


